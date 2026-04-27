"""
EU Funding & Tenders Portal – Project + Beneficiaries Scraper
=============================================================
Scrapes EU-funded research projects related to data spaces and related
topics using two sources:

  1. OpenAIRE REST API  → project metadata (grant ID, title, dates, funding)
  2. CORDIS bulk CSV    → complete beneficiary / organisation lists

Outputs (saved in the same folder as this script):
  - eu_projects_data_spaces.csv        (one row per project)
  - eu_beneficiaries_data_spaces.csv   (one row per organisation per project)
  - eu_projects_data_spaces.xlsx       (three sheets: Projects, Beneficiaries, Summaries)

Usage:
  python eu_funding_scraper.py

Requirements:
  pip install requests pandas openpyxl
"""

import io
import ssl
import time
import zipfile
import warnings
import requests
import pandas as pd
from pathlib import Path
from datetime import datetime

try:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except ImportError:
    pass

# ─────────────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────────────

SEARCH_TERMS = [
    "data spaces",
    "data space",
    "GAIA-X",
    "common european data space",
    "federated data",
    "data marketplace",
    "data sharing infrastructure",
    "data governance AI",
    "data lab european",
    "data interoperability",
    "data ecosystem",
    "data economy",
    "AI data infrastructure",
    "digital twin data",
    "smart data",
]

# OpenAIRE API
OPENAIRE_BASE  = "https://api.openaire.eu/search/projects"
FUNDER         = "EC"          # European Commission only; set "" for all
PAGE_SIZE      = 200
REQUEST_DELAY  = 1.0           # seconds between API calls

# CORDIS bulk-download URLs (contain both project.csv + organization.csv)
CORDIS_SOURCES = {
    "HORIZON": "https://cordis.europa.eu/data/cordis-HORIZONprojects-csv.zip",
    "H2020":   "https://cordis.europa.eu/data/cordis-h2020projects-csv.zip",
}

# Set False if SSL errors appear behind a corporate proxy
SSL_VERIFY = True

OUTPUT_DIR = Path(__file__).parent

# ─────────────────────────────────────────────────────────────────────────────
# OpenAIRE helpers
# ─────────────────────────────────────────────────────────────────────────────

def _leaf(obj, default=""):
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get("$", default)
    if isinstance(obj, list):
        return obj[0].get("$", default) if obj and isinstance(obj[0], dict) else default
    return str(obj)


def _first(obj):
    return (obj[0] if obj else None) if isinstance(obj, list) else obj


def parse_funding(ft_raw):
    """Return (funder, programme_code, programme_desc, action_type, action_desc)."""
    ft = _first(ft_raw)
    if not isinstance(ft, dict):
        return ("", "", "", "", "")
    funder = _leaf(ft.get("funder", {}).get("name"))
    fl1 = ft.get("funding_level_1", {})
    action = _leaf(fl1.get("name"))
    action_desc = _leaf(fl1.get("description"))
    # funding_level_0 is nested inside funding_level_1.parent
    fl0 = fl1.get("parent", {}).get("funding_level_0", {})
    prog = _leaf(fl0.get("name"))
    prog_desc = _leaf(fl0.get("description"))
    if not prog:                          # fallback: older records
        fl0b = ft.get("funding_level_0", {})
        prog      = _leaf(fl0b.get("name"))
        prog_desc = _leaf(fl0b.get("description"))
    return funder, prog, prog_desc, action, action_desc


def parse_project(result: dict) -> dict | None:
    try:
        proj = (result
                .get("metadata", {})
                .get("oaf:entity", {})
                .get("oaf:project", {}))
        if not proj:
            return None
        grant_id = _leaf(proj.get("code"))
        if not grant_id:
            return None
        funder, prog, prog_desc, action, action_desc = parse_funding(
            proj.get("fundingtree", {}))
        subjects_raw = proj.get("subject", [])
        if isinstance(subjects_raw, dict):
            subjects_raw = [subjects_raw]
        subjects = "; ".join(
            s.get("$", "") for s in (subjects_raw or [])
            if isinstance(s, dict) and s.get("$"))
        return {
            "grant_id":           grant_id,
            "acronym":            _leaf(proj.get("acronym")),
            "title":              _leaf(proj.get("title")),
            "start_date":         _leaf(proj.get("startdate")),
            "end_date":           _leaf(proj.get("enddate")),
            "duration_months":    _leaf(proj.get("duration")),
            "total_cost_eur":     _leaf(proj.get("totalcost")),
            "ec_contribution_eur":_leaf(proj.get("fundedamount")),
            "funder":             funder,
            "call_id":            _leaf(proj.get("callidentifier")),
            "programme":          prog,
            "programme_desc":     prog_desc,
            "action_type":        action,
            "action_type_desc":   action_desc,
            "contract_type":      _leaf(proj.get("contracttype")),
            "subjects":           subjects,
            "summary":            str(_leaf(proj.get("summary")))[:2000],
            "cordis_url":         f"https://cordis.europa.eu/project/id/{grant_id}",
        }
    except Exception as e:
        print(f"    [WARN] parse_project: {e}")
        return None


def fetch_projects(keyword: str) -> list[dict]:
    print(f"  Searching: '{keyword}'")
    all_projects, page = [], 1
    while True:
        params = {"keywords": keyword, "format": "json",
                  "size": PAGE_SIZE, "page": page}
        if FUNDER:
            params["funder"] = FUNDER
        try:
            resp = requests.get(OPENAIRE_BASE, params=params,
                                timeout=30, verify=SSL_VERIFY)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            print(f"    [ERROR] page {page}: {e}")
            break
        header  = data.get("response", {}).get("header", {})
        total   = int(header.get("total", {}).get("$", 0))
        results = data.get("response", {}).get("results", {}).get("result", []) or []
        if page == 1:
            print(f"    → {total} total results")
        if not results:
            break
        for r in results:
            p = parse_project(r)
            if p:
                all_projects.append(p)
        if page * PAGE_SIZE >= total:
            break
        page += 1
        time.sleep(REQUEST_DELAY)
    return all_projects


# ─────────────────────────────────────────────────────────────────────────────
# CORDIS bulk-download helpers
# ─────────────────────────────────────────────────────────────────────────────

def download_cordis_zip(label: str, url: str) -> zipfile.ZipFile | None:
    """Download a CORDIS zip and return as an in-memory ZipFile."""
    print(f"  Downloading CORDIS {label} bulk data …")
    try:
        resp = requests.get(url, timeout=120, stream=True, verify=SSL_VERIFY)
        resp.raise_for_status()
        total_mb = int(resp.headers.get("Content-Length", 0)) / 1024 / 1024
        chunks = []
        downloaded = 0
        for chunk in resp.iter_content(chunk_size=1024 * 1024):
            chunks.append(chunk)
            downloaded += len(chunk)
            pct = downloaded / (total_mb * 1024 * 1024) * 100 if total_mb else 0
            print(f"\r    {downloaded/1024/1024:.1f}/{total_mb:.1f} MB ({pct:.0f}%)",
                  end="", flush=True)
        print()
        zf = zipfile.ZipFile(io.BytesIO(b"".join(chunks)))
        print(f"    ✓ Files inside: {zf.namelist()}")
        return zf
    except Exception as e:
        print(f"    [ERROR] Could not download {label}: {e}")
        return None


def read_csv_from_zip(zf: zipfile.ZipFile, filename: str) -> pd.DataFrame | None:
    """Read a named CSV from an in-memory ZipFile."""
    if filename not in zf.namelist():
        print(f"    [WARN] '{filename}' not found in zip")
        return None
    try:
        with zf.open(filename) as f:
            content = f.read().decode("utf-8", errors="replace")
        return pd.read_csv(
            io.StringIO(content), sep=";",
            dtype=str, on_bad_lines="skip", low_memory=False
        )
    except Exception as e:
        print(f"    [ERROR] reading {filename}: {e}")
        return None


def get_cordis_beneficiaries(grant_ids: set[str]) -> pd.DataFrame:
    """
    Download CORDIS organisation.csv files for HORIZON and H2020,
    filter to our matched grant IDs, and return a combined DataFrame.
    """
    org_frames = []
    proj_frames = []

    for label, url in CORDIS_SOURCES.items():
        zf = download_cordis_zip(label, url)
        if zf is None:
            continue

        # ── organisations ──────────────────────────────────────────────────
        orgs = read_csv_from_zip(zf, "organization.csv")
        if orgs is not None:
            orgs["source_programme"] = label
            # projectID in CORDIS matches grant_id from OpenAIRE
            orgs["projectID"] = orgs["projectID"].astype(str).str.strip()
            matched = orgs[orgs["projectID"].isin(grant_ids)].copy()
            print(f"    {label} orgs: {len(orgs):,} total → {len(matched):,} matched")
            org_frames.append(matched)

        # ── project objective from project.csv (richer than OpenAIRE) ──────
        projs = read_csv_from_zip(zf, "project.csv")
        if projs is not None:
            projs["id"] = projs["id"].astype(str).str.strip()
            matched_p = projs[projs["id"].isin(grant_ids)][
                ["id", "objective", "keywords", "status", "frameworkProgramme",
                 "fundingScheme", "topics", "masterCall"]
            ].copy()
            matched_p.rename(columns={"id": "grant_id"}, inplace=True)
            proj_frames.append(matched_p)

    beneficiaries = pd.concat(org_frames, ignore_index=True) if org_frames else pd.DataFrame()
    cordis_extras  = pd.concat(proj_frames, ignore_index=True) if proj_frames else pd.DataFrame()

    return beneficiaries, cordis_extras


# ─────────────────────────────────────────────────────────────────────────────
# Excel formatting
# ─────────────────────────────────────────────────────────────────────────────

HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"
ALT_BG    = "EBF3FB"


def style_sheet(ws, col_widths: dict, wrap_cols: list[int] = None,
                number_cols: list[int] = None):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    hfont = Font(color=HEADER_FG, bold=True)
    hfill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    afill = PatternFill(start_color=ALT_BG,    end_color=ALT_BG,    fill_type="solid")

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for cell in ws[1]:
        cell.font      = hfont
        cell.fill      = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes   = "A2"
    ws.auto_filter.ref = ws.dimensions

    wrap_set   = set(wrap_cols or [])
    number_set = set(number_cols or [])

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = afill if row_idx % 2 == 0 else None
        for cell in row:
            wrap = cell.column in wrap_set
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)
            if fill:
                cell.fill = fill
            if cell.column in number_set:
                cell.number_format = "#,##0.00"


def write_excel(xlsx_path: Path, df_proj: pd.DataFrame,
                df_ben: pd.DataFrame, df_sum: pd.DataFrame):
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:

        # ── Sheet 1: Projects ──────────────────────────────────────────────
        proj_cols = [
            "grant_id", "acronym", "title", "start_date", "end_date",
            "duration_months", "total_cost_eur", "ec_contribution_eur",
            "funder", "call_id", "programme", "programme_desc",
            "action_type", "action_type_desc", "contract_type",
            "topics", "masterCall", "fundingScheme", "status",
            "subjects", "cordis_url",
        ]
        proj_cols = [c for c in proj_cols if c in df_proj.columns]
        df_proj[proj_cols].to_excel(writer, sheet_name="Projects", index=False)
        style_sheet(writer.sheets["Projects"], {
            1: 14, 2: 16, 3: 55, 4: 12, 5: 12, 6: 10,
            7: 18, 8: 20, 9: 22, 10: 38, 11: 14, 12: 38,
            13: 18, 14: 40, 15: 18, 16: 38, 17: 30, 18: 20,
            19: 12, 20: 45, 21: 50,
        }, wrap_cols=[3, 20], number_cols=[7, 8])

        # ── Sheet 2: Beneficiaries ─────────────────────────────────────────
        if not df_ben.empty:
            ben_cols = [
                "projectID", "projectAcronym", "source_programme",
                "name", "shortName", "role",
                "country", "city", "nutsCode",
                "activityType", "SME",
                "ecContribution", "netEcContribution", "totalCost",
                "organisationID", "vatNumber",
                "organizationURL",
                "active", "endOfParticipation",
            ]
            ben_cols = [c for c in ben_cols if c in df_ben.columns]
            df_ben[ben_cols].to_excel(writer, sheet_name="Beneficiaries", index=False)
            style_sheet(writer.sheets["Beneficiaries"], {
                1: 14, 2: 16, 3: 14, 4: 45, 5: 18, 6: 18,
                7: 12, 8: 20, 9: 12, 10: 20, 11: 8,
                12: 18, 13: 18, 14: 18,
                15: 16, 16: 16, 17: 40, 18: 10, 19: 20,
            }, wrap_cols=[4], number_cols=[12, 13, 14])

        # ── Sheet 3: Summaries / Objectives ───────────────────────────────
        sum_cols = [c for c in ["grant_id", "acronym", "title",
                                "programme", "objective", "summary"]
                    if c in df_sum.columns]
        df_sum[sum_cols].to_excel(writer, sheet_name="Summaries", index=False)
        style_sheet(writer.sheets["Summaries"], {
            1: 14, 2: 16, 3: 55, 4: 14, 5: 100, 6: 100,
        }, wrap_cols=[5, 6])

    print(f"✓ Excel saved: {xlsx_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  EU Funding Scraper – Data Spaces & Related Topics")
    print(f"  Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 65)
    print(f"  Source 1: OpenAIRE API  (project metadata)")
    print(f"  Source 2: CORDIS bulk   (beneficiary organisations)")
    print(f"  Funder:   {FUNDER or 'All'}")
    print(f"  Terms:    {len(SEARCH_TERMS)} search terms\n")

    # ── Step 1: Collect projects via OpenAIRE ──────────────────────────────
    print("─" * 65)
    print("STEP 1 – Querying OpenAIRE for EU-funded projects")
    print("─" * 65)
    all_projects: dict[str, dict] = {}
    for term in SEARCH_TERMS:
        for p in fetch_projects(term):
            if p["grant_id"] not in all_projects:
                all_projects[p["grant_id"]] = p
        print(f"    ✓ Running total: {len(all_projects)} unique projects\n")
        time.sleep(REQUEST_DELAY)

    if not all_projects:
        print("No projects found. Exiting.")
        return

    grant_ids = set(all_projects.keys())
    print(f"\n  Total unique projects: {len(grant_ids)}")

    # ── Step 2: Enrich from CORDIS bulk data ──────────────────────────────
    print("\n" + "─" * 65)
    print("STEP 2 – Downloading CORDIS bulk data for beneficiaries")
    print("─" * 65)
    df_ben, df_cordis_extras = get_cordis_beneficiaries(grant_ids)

    # ── Step 3: Build project DataFrame, merge CORDIS extras ──────────────
    df_proj = pd.DataFrame(list(all_projects.values()))
    if not df_cordis_extras.empty:
        df_proj = df_proj.merge(df_cordis_extras, on="grant_id", how="left")

    # Sort by start date
    df_proj["start_date"] = pd.to_datetime(df_proj["start_date"], errors="coerce")
    df_proj = df_proj.sort_values("start_date", ascending=False)
    df_proj["start_date"] = df_proj["start_date"].dt.strftime("%Y-%m-%d").fillna("")
    for col in ["total_cost_eur", "ec_contribution_eur"]:
        df_proj[col] = pd.to_numeric(df_proj[col], errors="coerce").fillna(0.0)

    # ── Step 4: Numeric beneficiary cost columns ───────────────────────────
    if not df_ben.empty:
        for col in ["ecContribution", "netEcContribution", "totalCost"]:
            if col in df_ben.columns:
                df_ben[col] = pd.to_numeric(df_ben[col], errors="coerce").fillna(0.0)

    # ── Step 5: Summaries sheet (merge CORDIS objective + OpenAIRE summary)
    df_sum = df_proj[["grant_id", "acronym", "title", "programme", "summary"]].copy()
    if "objective" in df_proj.columns:
        df_sum["objective"] = df_proj["objective"]

    # ── Step 6: Save outputs ───────────────────────────────────────────────
    print("\n" + "─" * 65)
    print("STEP 3 – Saving outputs")
    print("─" * 65)

    csv_proj = OUTPUT_DIR / "eu_projects_data_spaces.csv"
    df_proj.to_csv(csv_proj, index=False, encoding="utf-8-sig")
    print(f"✓ Projects CSV:      {csv_proj.name}  ({len(df_proj)} rows)")

    if not df_ben.empty:
        csv_ben = OUTPUT_DIR / "eu_beneficiaries_data_spaces.csv"
        df_ben.to_csv(csv_ben, index=False, encoding="utf-8-sig")
        print(f"✓ Beneficiaries CSV: {csv_ben.name}  ({len(df_ben)} rows)")

    xlsx_path = OUTPUT_DIR / "eu_projects_data_spaces.xlsx"
    write_excel(xlsx_path, df_proj, df_ben, df_sum)

    # ── Summary stats ──────────────────────────────────────────────────────
    print("\n" + "─" * 65)
    print("  SUMMARY")
    print("─" * 65)
    print(f"  Projects found:         {len(df_proj)}")

    valid = pd.to_datetime(df_proj["start_date"], errors="coerce").dropna()
    if not valid.empty:
        print(f"  Year range:             {int(valid.dt.year.min())} – {int(valid.dt.year.max())}")
    ec = df_proj["ec_contribution_eur"].sum()
    if ec > 0:
        print(f"  Total EC contribution:  €{ec:,.0f}")

    print(f"\n  Top programmes:")
    for (prog, desc), cnt in (
        df_proj.groupby(["programme", "programme_desc"])
               .size().sort_values(ascending=False).head(5).items()
    ):
        print(f"    • {prog} – {desc}: {cnt}")

    if not df_ben.empty:
        print(f"\n  Beneficiaries total:    {len(df_ben)}")
        print(f"  Unique organisations:   {df_ben['name'].nunique()}")
        print(f"  Countries represented:  {df_ben['country'].nunique()}")
        print(f"\n  Top 10 countries by number of participations:")
        for country, cnt in df_ben["country"].value_counts().head(10).items():
            print(f"    • {country}: {cnt}")
        print(f"\n  Roles breakdown:")
        for role, cnt in df_ben["role"].value_counts().items():
            print(f"    • {role}: {cnt}")

    print("─" * 65)
    print(f"\n  Done! {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
